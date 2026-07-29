"""
受限 Python 执行环境：用代码完成复杂文档排版/编辑。

允许：docx / openpyxl / pypdf / reportlab / pathlib / json 等
禁止：网络、子进程、任意写系统盘（输出目录限制在 data/docs_out）
"""
from __future__ import annotations

import ast
import traceback
from pathlib import Path
from typing import Any

from agent.doc_ops import docs_out_dir, resolve_doc_path
from agent.llm_client import app_dir

_FORBIDDEN_NAMES = {
    "eval",
    "exec",
    "compile",
    "__import__",
    "input",
    "breakpoint",
}

_FORBIDDEN_ATTR = {
    "system",
    "popen",
    "Popen",
    "remove",
    "unlink",
    "rmdir",
    "removedirs",
}


def _validate_code(code: str) -> None:
    try:
        tree = ast.parse(code)
    except SyntaxError as e:
        raise ValueError(f"脚本语法错误: {e}") from e
    for node in ast.walk(tree):
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            # 只允许白名单模块
            names = []
            if isinstance(node, ast.Import):
                names = [a.name.split(".")[0] for a in node.names]
            else:
                names = [node.module.split(".")[0]] if node.module else []
            for n in names:
                if n not in _ALLOWED_ROOTS:
                    raise ValueError(f"不允许 import {n}（仅限文档相关库）")
        if isinstance(node, ast.Call):
            if isinstance(node.func, ast.Name) and node.func.id in _FORBIDDEN_NAMES:
                raise ValueError(f"不允许调用 {node.func.id}()")
        if isinstance(node, ast.Attribute) and node.attr in _FORBIDDEN_ATTR:
            raise ValueError(f"不允许属性/方法 .{node.attr}")


_ALLOWED_ROOTS = {
    "docx",
    "openpyxl",
    "pypdf",
    "reportlab",
    "pathlib",
    "json",
    "re",
    "copy",
    "datetime",
    "math",
    "collections",
    "typing",
    "io",
    "base64",
    "zipfile",
}


def _under(child: Path, parent: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def _safe_open(file, mode="r", *args, **kwargs):
    path = Path(file).expanduser().resolve()
    out_root = docs_out_dir().resolve()
    app_root = app_dir().resolve()
    writing = any(x in mode for x in ("w", "a", "x", "+"))
    if writing:
        if not _under(path, out_root):
            raise PermissionError(f"只允许写入 {out_root}，拒绝: {path}")
    else:
        if not (_under(path, app_root) or _under(path, out_root)):
            raise PermissionError(f"只允许读取项目目录内文件: {path}")
    return open(path, mode, *args, **kwargs)  # noqa: SIM115


def run_document_code(code: str, *, input_files_json: str = "") -> str:
    """
    执行用户/Agent 生成的排版脚本。
    预置变量:
      OUT_DIR, APP_DIR, INPUT_FILES(list[Path]), resolve_path, print 收集到日志
    """
    code = (code or "").strip()
    if not code:
        return "代码为空"
    if len(code) > 20000:
        return "代码过长（>20KB）"
    _validate_code(code)

    input_files: list[Path] = []
    if input_files_json.strip():
        import json

        raw = json.loads(input_files_json)
        if isinstance(raw, list):
            for item in raw:
                input_files.append(resolve_doc_path(str(item)))

    logs: list[str] = []

    def _log(*args, **kwargs):
        logs.append(" ".join(str(a) for a in args))

    # 预导入允许库
    import copy
    import datetime
    import json
    import math
    import re
    from pathlib import Path as PathCls

    try:
        import docx
        from docx import Document
        from docx.enum.text import WD_ALIGN_PARAGRAPH
        from docx.shared import Cm, Inches, Pt, RGBColor
    except ImportError:
        docx = Document = WD_ALIGN_PARAGRAPH = Cm = Inches = Pt = RGBColor = None  # type: ignore

    try:
        import openpyxl
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        openpyxl = Workbook = load_workbook = Font = Alignment = PatternFill = None  # type: ignore

    try:
        import pypdf
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        pypdf = PdfReader = PdfWriter = None  # type: ignore

    try:
        import reportlab
        from reportlab.lib.pagesizes import A4, letter
        from reportlab.pdfgen import canvas
    except ImportError:
        reportlab = A4 = letter = canvas = None  # type: ignore

    glb: dict[str, Any] = {
        "__builtins__": {
            "abs": abs,
            "min": min,
            "max": max,
            "sum": sum,
            "len": len,
            "range": range,
            "enumerate": enumerate,
            "zip": zip,
            "list": list,
            "dict": dict,
            "str": str,
            "int": int,
            "float": float,
            "bool": bool,
            "True": True,
            "False": False,
            "None": None,
            "print": _log,
            "open": _safe_open,
            "isinstance": isinstance,
            "hasattr": hasattr,
            "getattr": getattr,
            "setattr": setattr,
            "sorted": sorted,
            "round": round,
            "Exception": Exception,
            "ValueError": ValueError,
            "TypeError": TypeError,
            "KeyError": KeyError,
        },
        "OUT_DIR": docs_out_dir(),
        "APP_DIR": app_dir(),
        "INPUT_FILES": input_files,
        "resolve_path": resolve_doc_path,
        "Path": PathCls,
        "json": json,
        "re": re,
        "copy": copy,
        "datetime": datetime,
        "math": math,
        "docx": docx,
        "Document": Document,
        "Pt": Pt,
        "Cm": Cm,
        "Inches": Inches,
        "RGBColor": RGBColor,
        "WD_ALIGN_PARAGRAPH": WD_ALIGN_PARAGRAPH,
        "openpyxl": openpyxl,
        "Workbook": Workbook,
        "load_workbook": load_workbook,
        "Font": Font,
        "Alignment": Alignment,
        "PatternFill": PatternFill,
        "pypdf": pypdf,
        "PdfReader": PdfReader,
        "PdfWriter": PdfWriter,
        "reportlab": reportlab,
        "A4": A4,
        "letter": letter,
        "canvas": canvas,
    }

    loc: dict[str, Any] = {}
    try:
        compiled = compile(code, "<doc_script>", "exec")
        exec(compiled, glb, loc)  # noqa: S102
    except Exception:
        return "执行失败:\n" + traceback.format_exc(limit=8)

    result = loc.get("RESULT") or glb.get("RESULT")
    msg = "脚本执行完成。"
    if result is not None:
        msg += f"\nRESULT={result}"
    if logs:
        msg += "\n--- 输出 ---\n" + "\n".join(logs[-50:])
    msg += f"\n输出目录: {docs_out_dir()}"
    return msg
